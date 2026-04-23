import io
import json
import re
import uuid
from dataclasses import dataclass
from typing import Iterable

from openpyxl import Workbook, load_workbook
from sqlalchemy import func

from .models import ConversionBatch, ProxyRecord, RelayServer, db


COMMON_COUNTRIES = [
    "美国",
    "英国",
    "日本",
    "德国",
    "法国",
    "新加坡",
    "韩国",
    "加拿大",
    "澳大利亚",
    "中国香港",
    "中国台湾",
    "阿联酋",
    "巴西",
    "意大利",
    "西班牙",
    "荷兰",
    "瑞士",
    "马来西亚",
    "印尼",
    "越南",
    "泰国",
    "印度",
    "墨西哥",
]

LINE_PATTERN = re.compile(
    r"^\s*(?P<host>[A-Za-z0-9._-]+)\s*:\s*(?P<port>\d{1,5})\s*:\s*"
    r"(?P<username>[^:\s{}]+)\s*:\s*(?P<password>[^{}]+?)\s*"
    r"(?:\{(?P<remark>[^}]*)\})?\s*$"
)


@dataclass
class ParsedProxy:
    original_line: str
    origin_host: str
    origin_port: int
    username: str
    password: str
    remark: str


@dataclass
class ConversionPayload:
    source_type: str
    country: str
    relay_server: RelayServer
    port_range_start: int
    port_range_end: int
    assigned_start_port: int
    assigned_end_port: int
    raw_input: str
    result_text: str
    result_json: str
    batch_code: str
    remark_prefix: str
    total_lines: int
    success_count: int
    skipped_count: int
    error_summary: str
    excel_rows: list[dict]
    json_items: list[dict]
    records: list[dict]
    errors: list[str]
    zero_summary: dict


class ConversionError(ValueError):
    pass


def extract_lines_from_text(raw_text: str) -> list[str]:
    return [line.strip() for line in raw_text.splitlines() if line.strip()]


def extract_lines_from_excel(file_storage) -> list[str]:
    workbook = load_workbook(file_storage, read_only=True, data_only=True)
    worksheet = workbook.active
    lines: list[str] = []

    for row in worksheet.iter_rows(values_only=True):
        values = [str(value).strip() for value in row if value is not None and str(value).strip()]
        if values:
            lines.append(" ".join(values))

    return lines


def parse_proxy_line(line: str, fallback_remark: str) -> ParsedProxy:
    match = LINE_PATTERN.match(line)
    if not match:
        raise ConversionError(f"无法识别的行格式: {line}")

    port = int(match.group("port"))
    if not 1 <= port <= 65535:
        raise ConversionError(f"原始端口不合法: {line}")

    remark = (match.group("remark") or fallback_remark or "").strip()
    if not remark:
        raise ConversionError(f"缺少备注信息: {line}")

    return ParsedProxy(
        original_line=line,
        origin_host=match.group("host").strip(),
        origin_port=port,
        username=match.group("username").strip(),
        password=match.group("password").strip(),
        remark=remark,
    )


def find_next_available_start_port(relay_server: RelayServer, required_count: int = 1) -> int | None:
    if required_count < 1:
        required_count = 1

    used_ports = {
        port
        for (port,) in (
            db.session.query(ProxyRecord.assigned_port)
            .filter(
                ProxyRecord.relay_server_id == relay_server.id,
                ProxyRecord.assigned_port >= relay_server.port_range_start,
                ProxyRecord.assigned_port <= relay_server.port_range_end,
            )
            .all()
        )
    }

    start = relay_server.port_range_start
    end = relay_server.port_range_end
    latest_candidate = end - required_count + 1

    # This favors straightforward behavior over a more complex interval structure.
    for candidate in range(start, latest_candidate + 1):
        if all((candidate + offset) not in used_ports for offset in range(required_count)):
            return candidate

    return None


def validate_port_block(relay_server: RelayServer, start_port: int, required_count: int) -> int:
    end_port = start_port + required_count - 1
    if start_port < relay_server.port_range_start or end_port > relay_server.port_range_end:
        raise ConversionError(
            f"端口范围超出中转服务器限制，可用范围为 {relay_server.port_range_start}-{relay_server.port_range_end}。"
        )

    conflicting = (
        db.session.query(ProxyRecord.assigned_port)
        .filter(
            ProxyRecord.relay_server_id == relay_server.id,
            ProxyRecord.assigned_port >= start_port,
            ProxyRecord.assigned_port <= end_port,
        )
        .order_by(ProxyRecord.assigned_port.asc())
        .first()
    )
    if conflicting:
        raise ConversionError(f"端口 {conflicting[0]} 已被占用，请改用自动分配或调整起始端口。")

    return end_port


def build_conversion_payload(
    *,
    source_type: str,
    lines: list[str],
    relay_server: RelayServer,
    country: str,
    remark_prefix: str,
    remark_start: int,
    manual_start_port: int | None,
) -> ConversionPayload:
    parsed_entries: list[ParsedProxy] = []
    errors: list[str] = []

    for index, line in enumerate(lines):
        fallback_remark = f"{remark_prefix}{remark_start + index}" if remark_prefix else f"Node_{remark_start + index}"
        try:
            parsed_entries.append(parse_proxy_line(line, fallback_remark=fallback_remark))
        except ConversionError as exc:
            errors.append(str(exc))

    if not parsed_entries:
        raise ConversionError("没有解析出有效代理，请检查导入内容格式。")

    required_count = len(parsed_entries)
    if manual_start_port:
        assigned_start_port = manual_start_port
    else:
        assigned_start_port = find_next_available_start_port(relay_server, required_count)
        if assigned_start_port is None:
            raise ConversionError("当前中转服务器端口范围内没有足够的连续可用端口。")

    assigned_end_port = validate_port_block(relay_server, assigned_start_port, required_count)

    json_items: list[dict] = []
    excel_rows: list[dict] = []
    record_payloads: list[dict] = []
    output_lines: list[str] = []

    current_port = assigned_start_port
    for entry in parsed_entries:
        json_item = {
            "dest": [f"{entry.origin_host}:{entry.origin_port}"],
            "listen_port": current_port,
            "name": entry.remark,
            "enable_udp": True,
        }
        forward_line = f"{relay_server.host}:{current_port}:{entry.username}:{entry.password}{{{entry.remark}}}"
        origin_line = f"{entry.origin_host}:{entry.origin_port}:{entry.username}:{entry.password}"

        json_items.append(json_item)
        output_lines.append(forward_line)
        excel_rows.append(
            {
                "国家": country,
                "备注": entry.remark,
                "交付格式": forward_line,
                "原始格式": origin_line,
                "中转服务器": relay_server.name,
                "分配端口": current_port,
            }
        )
        record_payloads.append(
            {
                "country": country,
                "remark": entry.remark,
                "origin_host": entry.origin_host,
                "origin_port": entry.origin_port,
                "username": entry.username,
                "password": entry.password,
                "assigned_port": current_port,
                "zero_port_id": None,
                "zero_sync_status": "pending",
                "zero_sync_error": "",
                "forward_line": forward_line,
                "origin_line": origin_line,
                "json_entry": json.dumps(json_item, ensure_ascii=False),
            }
        )
        current_port += 1

    return ConversionPayload(
        source_type=source_type,
        country=country,
        relay_server=relay_server,
        port_range_start=relay_server.port_range_start,
        port_range_end=relay_server.port_range_end,
        assigned_start_port=assigned_start_port,
        assigned_end_port=assigned_end_port,
        raw_input="\n".join(lines),
        result_text="\n".join(output_lines),
        result_json=json.dumps(json_items, ensure_ascii=False, indent=2),
        batch_code=f"BATCH-{uuid.uuid4().hex[:10].upper()}",
        remark_prefix=remark_prefix,
        total_lines=len(lines),
        success_count=len(parsed_entries),
        skipped_count=len(lines) - len(parsed_entries),
        error_summary="\n".join(errors),
        excel_rows=excel_rows,
        json_items=json_items,
        records=record_payloads,
        errors=errors,
        zero_summary={"enabled": False, "ok": 0, "failed": 0, "dry_run": False},
    )


def sync_payload_render_fields(payload: ConversionPayload) -> None:
    payload.assigned_start_port = min(record["assigned_port"] for record in payload.records)
    payload.assigned_end_port = max(record["assigned_port"] for record in payload.records)
    payload.result_text = "\n".join(record["forward_line"] for record in payload.records)
    payload.result_json = json.dumps(
        [json.loads(record["json_entry"]) for record in payload.records],
        ensure_ascii=False,
        indent=2,
    )
    payload.excel_rows = [
        {
            "国家": record["country"],
            "备注": record["remark"],
            "交付格式": record["forward_line"],
            "原始格式": record["origin_line"],
            "中转服务器": payload.relay_server.name,
            "分配端口": record["assigned_port"],
        }
        for record in payload.records
    ]


def persist_conversion(payload: ConversionPayload) -> ConversionBatch:
    batch = ConversionBatch(
        batch_code=payload.batch_code,
        source_type=payload.source_type,
        relay_server_id=payload.relay_server.id,
        relay_name=payload.relay_server.name,
        relay_host=payload.relay_server.host,
        port_range_start=payload.port_range_start,
        port_range_end=payload.port_range_end,
        assigned_start_port=payload.assigned_start_port,
        assigned_end_port=payload.assigned_end_port,
        country=payload.country,
        total_lines=payload.total_lines,
        success_count=payload.success_count,
        skipped_count=payload.skipped_count,
        remark_prefix=payload.remark_prefix,
        raw_input=payload.raw_input,
        result_text=payload.result_text,
        result_json=payload.result_json,
        error_summary=payload.error_summary,
    )
    db.session.add(batch)
    db.session.flush()

    for record in payload.records:
        db.session.add(
            ProxyRecord(
                batch_id=batch.id,
                relay_server_id=payload.relay_server.id,
                **record,
            )
        )

    db.session.commit()
    return batch


def build_excel_bytes(rows: Iterable[dict]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "delivery"

    rows_list = list(rows)
    if not rows_list:
        rows_list = [{"备注": "", "交付格式": "", "原始格式": ""}]

    headers = list(rows_list[0].keys())
    worksheet.append(headers)
    for row in rows_list:
        worksheet.append([row.get(header, "") for header in headers])

    for column_letter, width in {
        "A": 14,
        "B": 20,
        "C": 54,
        "D": 54,
        "E": 18,
        "F": 12,
    }.items():
        worksheet.column_dimensions[column_letter].width = width

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_excel_rows_from_batch(batch: ConversionBatch) -> list[dict]:
    return [
        {
            "国家": record.country,
            "备注": record.remark,
            "交付格式": record.forward_line,
            "原始格式": record.origin_line,
            "中转服务器": batch.relay_name,
            "分配端口": record.assigned_port,
        }
        for record in batch.records
    ]


def build_country_remark_prefix(country: str) -> str:
    return f"{country}电商" if country else "Proxy-"


def get_country_profile(country: str) -> dict:
    country = (country or "").strip()
    default_prefix = build_country_remark_prefix(country)
    if not country:
        return {
            "country": "",
            "suggested_prefix": "Proxy-",
            "next_number": 1,
            "record_count": 0,
            "latest_remark": "",
        }

    records = (
        ProxyRecord.query.filter_by(country=country)
        .order_by(ProxyRecord.created_at.desc(), ProxyRecord.id.desc())
        .all()
    )

    max_number = 0
    latest_remark = records[0].remark if records else ""
    pattern = re.compile(rf"^{re.escape(default_prefix)}\s*(\d+)$")

    for record in records:
        matched = pattern.match(record.remark.strip())
        if matched:
            max_number = max(max_number, int(matched.group(1)))

    return {
        "country": country,
        "suggested_prefix": default_prefix,
        "next_number": max_number + 1 if max_number else 1,
        "record_count": len(records),
        "latest_remark": latest_remark,
    }


def get_dashboard_stats() -> dict:
    relay_count = db.session.query(func.count(RelayServer.id)).filter(RelayServer.is_active.is_(True)).scalar() or 0
    batch_count = db.session.query(func.count(ConversionBatch.id)).scalar() or 0
    proxy_count = db.session.query(func.count(ProxyRecord.id)).scalar() or 0
    country_count = db.session.query(func.count(func.distinct(ProxyRecord.country))).scalar() or 0
    zero_proxy_count = (
        db.session.query(func.count(ProxyRecord.id))
        .filter(ProxyRecord.zero_sync_status == "ok")
        .scalar()
        or 0
    )

    return {
        "relay_count": relay_count,
        "batch_count": batch_count,
        "proxy_count": proxy_count,
        "country_count": country_count,
        "zero_proxy_count": zero_proxy_count,
    }


def get_country_groups(limit: int = 12) -> list[dict]:
    groups = (
        db.session.query(
            ProxyRecord.country.label("country"),
            func.count(ProxyRecord.id).label("count"),
            func.max(ProxyRecord.created_at).label("latest_at"),
        )
        .group_by(ProxyRecord.country)
        .order_by(func.count(ProxyRecord.id).desc(), func.max(ProxyRecord.created_at).desc())
        .limit(limit)
        .all()
    )

    results: list[dict] = []
    for group in groups:
        recent_records = (
            ProxyRecord.query.filter_by(country=group.country)
            .order_by(ProxyRecord.created_at.desc())
            .limit(5)
            .all()
        )
        results.append(
            {
                "country": group.country,
                "count": group.count,
                "latest_at": group.latest_at,
                "records": recent_records,
            }
        )

    return results
